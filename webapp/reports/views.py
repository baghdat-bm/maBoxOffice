from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import permission_required
from django.core.paginator import Paginator
from datetime import date
# from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, Border, Side
from django.db.models import Count, Q, Sum, Case, When, F, Value, IntegerField, OuterRef, Subquery
from django.contrib import messages
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict

from references.models import Event, EventTemplate
from ticket_sales.models import TicketSalesTicket, TicketSalesPayments, TicketSalesService, TicketSale, SaleTypeEnum
from .forms import TicketReportForm, SalesReportForm, SessionsReportForm, ServiceReportForm


from .utils import get_filtered_sales_data, get_sessions_report_data, get_ticket_report_data, get_services_report_data


@permission_required('reports.view_sales_report', raise_exception=True)
def sales_report(request):
    # Получаем данные из формы фильтрации
    form = SalesReportForm(request.GET or None)

    if form.is_valid():

        tickets = get_filtered_sales_data(request, form.cleaned_data)

        # Пагинация по 20 записей на страницу
        paginator = Paginator(tickets, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Итоговые суммы по текущей странице
        totals = {
            'total_amount': sum(item['total_amount'] for item in page_obj),
            'total_cashier_paid_card': sum(item['cashier_paid_card'] for item in page_obj),
            'total_cashier_paid_cash': sum(item['cashier_paid_cash'] for item in page_obj),
            'total_kiosk_paid': sum(item['kiosk_paid'] for item in page_obj),
            'total_muzaidyny_qr_paid': sum(item['muzaidyny_qr_paid'] for item in page_obj),
            'total_muzaidyny_card_paid': sum(item['muzaidyny_card_paid'] for item in page_obj),
            'total_kaspi_paid': sum(item['kaspi_paid'] for item in page_obj),
            'total_refund_amount': sum(item['refund_amount'] for item in page_obj),
        }

        # Получаем дополнительные данные для контекста
        sale_types = request.GET.getlist('sale_types')
        events = request.GET.getlist('events')

        # Если не выбраны sale_types или events, ставим "all" по умолчанию
        if len(sale_types) == 0:
            sale_types = ['all', ]
        if len(events) == 0:
            events = ['all', ]

    else:
        page_obj = None  # Ошибки формы
        totals = {}  # Пустые итоги
        sale_types = []
        events = []
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')

    context = {
        'form': form,
        'page_obj': page_obj,
        'totals': totals,  # Передаем итоговые суммы в контекст
        'sale_type_choices': SaleTypeEnum.choices(),
        'events_list': Event.objects.all(),
        'selected_sale_types': sale_types,  # Передаем выбранные типы продаж
        'selected_events': events,  # Передаем выбранные мероприятия
    }
    return render(request, 'reports/sales_report.html', context)


@permission_required('reports.view_sales_report', raise_exception=True)
def sales_report_export(request):
    # Получаем данные из формы фильтрации
    form = SalesReportForm(request.GET or None)

    if form.is_valid():
        tickets = get_filtered_sales_data(request, form.cleaned_data)

        # Создаем Excel-файл
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales Report"

        # Заголовки для столбцов
        headers = [
            "№", "Дата", "Сумма продажи", "Касса (безнал)", "Касса (наличные)",
            "Киоск", "Muzaidyny.kz (KaspiQR)", "Muzaidyny.kz (Карта)",
            "Kaspi платежи", "Возвраты", "Мероприятие"
        ]
        sheet.append(headers)

        # Make the header row bold and add borders
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Variables to hold summary totals
        total_amount = 0
        total_cashier_card = 0
        total_cashier_cash = 0
        total_kiosk = 0
        total_muzaidyny_qr = 0
        total_muzaidyny_card = 0
        total_kaspi = 0
        total_refund = 0

        curr_no = 1
        # Заполняем строки данными отчета
        for ticket in tickets:
            row = [
                curr_no,
                ticket['sale_date'].strftime('%d.%m.%Y'),
                ticket['total_amount'],
                ticket['cashier_paid_card'],
                ticket['cashier_paid_cash'],
                ticket['kiosk_paid'],
                ticket['muzaidyny_qr_paid'],
                ticket['muzaidyny_card_paid'],
                ticket['kaspi_paid'],
                ticket['refund_amount'],
                ticket['event_name'],
            ]
            sheet.append(row)

            # Apply borders to each cell in the row
            for cell in sheet[sheet.max_row]:
                cell.border = thin_border

            # Accumulate totals
            total_amount += ticket['total_amount']
            total_cashier_card += ticket['cashier_paid_card']
            total_cashier_cash += ticket['cashier_paid_cash']
            total_kiosk += ticket['kiosk_paid']
            total_muzaidyny_qr += ticket['muzaidyny_qr_paid']
            total_muzaidyny_card += ticket['muzaidyny_card_paid']
            total_kaspi += ticket['kaspi_paid']
            total_refund += ticket['refund_amount']

            curr_no += 1

        # Добавляем итоговую строку
        total_row = [
            "Итого",
            "",  # Leave the date column empty for the summary row
            total_amount,
            total_cashier_card,
            total_cashier_cash,
            total_kiosk,
            total_muzaidyny_qr,
            total_muzaidyny_card,
            total_kaspi,
            total_refund,
            ""  # Leave the event name column empty for the summary row
        ]
        sheet.append(total_row)

        # Make the total row bold and add borders
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Apply borders to header row
        for cell in sheet[1]:
            cell.border = thin_border

        # Настраиваем ширину колонок
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 15

        # Создаем HTTP-ответ с Excel-файлом
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=sales_report_{now().strftime("%Y-%m-%d")}.xlsx'
        workbook.save(response)

        return response

    else:
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')
        return redirect('reports:sales_report')


@permission_required('reports.view_events_report', raise_exception=True)
def sessions_report(request):
    form = SessionsReportForm(request.GET or None)

    if form.is_valid():
        tickets_grouped = get_sessions_report_data(form)

        # Paginate the results
        paginator = Paginator(tickets_grouped, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calculate total summary for paginated results
        total_summary = {
            'total_quantity': sum(item['event_quantity'] for item in page_obj.object_list),
            'total_sold': sum(item['total_tickets_sold'] for item in page_obj.object_list),
            'total_left': sum(item['total_tickets_left'] for item in page_obj.object_list),
            'total_card_sales_cs': sum(item['total_card_sales_cs'] for item in page_obj.object_list),
            'total_cash_sales_cs': sum(item['total_cash_sales_cs'] for item in page_obj.object_list),
            'total_kiosk_sales': sum(item['total_kiosk_sales'] for item in page_obj.object_list),
            'total_qr_sales_sm': sum(item['total_qr_sales_sm'] for item in page_obj.object_list),
            'total_card_sales_sm': sum(item['total_card_sales_sm'] for item in page_obj.object_list),
            'total_kaspi_sales': sum(item['total_kaspi_sales'] for item in page_obj.object_list),
            'total_refunds': sum(item['total_refunds'] for item in page_obj.object_list),
        }

        events_list = EventTemplate.objects.all()
        selected_events = form.cleaned_data.get('event_templates', ['all'])
    else:
        page_obj = None
        total_summary = {}
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')

        events_list = EventTemplate.objects.all()
        selected_events = ['all']

    context = {
        'form': form,
        'page_obj': page_obj,
        'events_list': events_list,
        'selected_events': selected_events,
        'total_summary': total_summary,
    }

    return render(request, 'reports/sessions_report.html', context)


@permission_required('reports.view_events_report', raise_exception=True)
def export_sessions_report_to_excel(request):
    form = SessionsReportForm(request.GET or None)

    if form.is_valid():
        tickets_grouped = get_sessions_report_data(form)

        # Create a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Отчет по сеансам'

        # Define headers for Excel columns
        headers = [
            '№', 'Дата', 'Время сеанса', 'Кол-во всего билетов', 'Кол-во остатков билетов',
            'Кол-во проданных билетов', 'Касса (безнал)', 'Касса (наличные)', 'Киоск',
            'Muzaidyny.kz (KaspiQR)', 'Muzaidyny.kz (Карта)', 'Kaspi платежи', 'Возвраты', 'Мероприятие'
        ]

        # Write headers to Excel sheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Make headers bold
            cell.font = Font(bold=True)

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write data rows
        for row_num, ticket in enumerate(tickets_grouped, start=2):
            ws.cell(row=row_num, column=1, value=row_num-1)  # №
            ws.cell(row=row_num, column=2, value=ticket['sale_date'].strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=3, value=ticket['event_time'].strftime('%H:%M'))
            ws.cell(row=row_num, column=4, value=ticket['event_quantity'])
            ws.cell(row=row_num, column=5, value=ticket['total_tickets_left'])
            ws.cell(row=row_num, column=6, value=ticket['total_tickets_sold'])
            ws.cell(row=row_num, column=7, value=ticket['total_card_sales_cs'])
            ws.cell(row=row_num, column=8, value=ticket['total_cash_sales_cs'])
            ws.cell(row=row_num, column=9, value=ticket['total_kiosk_sales'])
            ws.cell(row=row_num, column=10, value=ticket['total_qr_sales_sm'])
            ws.cell(row=row_num, column=11, value=ticket['total_card_sales_sm'])
            ws.cell(row=row_num, column=12, value=ticket['total_kaspi_sales'])
            ws.cell(row=row_num, column=13, value=ticket['total_refunds'])
            ws.cell(row=row_num, column=14, value=ticket['event_name'])

            # Apply borders to each cell in the current row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

        # Calculate total summary for paginated results
        total_summary = {
            'total_quantity': sum(item['event_quantity'] for item in tickets_grouped),
            'total_sold': sum(item['total_tickets_sold'] for item in tickets_grouped),
            'total_left': sum(item['total_tickets_left'] for item in tickets_grouped),
            'total_card_sales_cs': sum(item['total_card_sales_cs'] for item in tickets_grouped),
            'total_cash_sales_cs': sum(item['total_cash_sales_cs'] for item in tickets_grouped),
            'total_kiosk_sales': sum(item['total_kiosk_sales'] for item in tickets_grouped),
            'total_qr_sales_sm': sum(item['total_qr_sales_sm'] for item in tickets_grouped),
            'total_card_sales_sm': sum(item['total_card_sales_sm'] for item in tickets_grouped),
            'total_kaspi_sales': sum(item['total_kaspi_sales'] for item in tickets_grouped),
            'total_refunds': sum(item['total_refunds'] for item in tickets_grouped),
        }

        # Add total summary row
        total_row = [
            'Итого',
            '', '',  # Skip date and time columns for the total row
            total_summary['total_quantity'],
            total_summary['total_left'],
            total_summary['total_sold'],
            total_summary['total_card_sales_cs'],
            total_summary['total_cash_sales_cs'],
            total_summary['total_kiosk_sales'],
            total_summary['total_qr_sales_sm'],
            total_summary['total_card_sales_sm'],
            total_summary['total_kaspi_sales'],
            total_summary['total_refunds'],
            ''  # Skip event name for the total row
        ]
        summary_row_num = len(tickets_grouped) + 2  # Place the summary row after all data rows
        for col_num, value in enumerate(total_row, 1):
            cell = ws.cell(row=summary_row_num, column=col_num, value=value)
            # Make the total row bold
            cell.font = Font(bold=True)
            # Apply borders to each cell in the total row
            cell.border = thin_border

        # Apply borders to the header row
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).border = thin_border

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create a response for the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sessions_report.xlsx"'
        wb.save(response)

        return response
    else:
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')
        return redirect('reports:sessions_report')


@permission_required('reports.view_tickets_report', raise_exception=True)
def tickets_report(request):
    form = TicketReportForm(request.GET or None)

    if form.is_valid():
        tickets = get_ticket_report_data(form)

        ticket_number = form.cleaned_data.get('ticket_number')
        order_number = form.cleaned_data.get('order_number')
        if (ticket_number or order_number) and not tickets.exists():
            if ticket_number:
                messages.warning(request, f'По номеру билета {ticket_number} данных не найдено')
            if order_number:
                messages.warning(request, f'По номеру заказа {order_number} данных не найдено')

        # Добавляем пагинацию
        paginator = Paginator(tickets, 20)  # 10 билетов на страницу
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calculate totals for all numeric fields
        total_summary = {
            'total_amount': sum(item.amount for item in page_obj.object_list),
        }
        events_list = EventTemplate.objects.all()
        selected_events = form.cleaned_data.get('event_templates', ['all'])

    else:
        page_obj = None  # Ошибки формы
        total_summary = {}
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')

        events_list = EventTemplate.objects.all()
        selected_events = ['all']

    context = {
        'form': form,
        'page_obj': page_obj,
        'events_list': events_list,
        'selected_events': selected_events,
        'total_summary': total_summary,
    }
    return render(request, 'reports/tickets_report.html', context)


@permission_required('reports.view_tickets_report', raise_exception=True)
def export_tickets_report_to_excel(request):
    form = TicketReportForm(request.GET or None)

    if form.is_valid():
        tickets = get_ticket_report_data(form)

        # Create a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Реестр билетов'

        # Define headers for Excel columns
        headers = [
            '№', '№ билета', '№ заказа', 'Дата и время сеанса', 'Стоимость билета',
            'Наименование услуги', 'Наименование инвентарья', 'Тип продажи',
            'Дата и время брони', 'Номер чек оплаты', 'Номер телефона клиента'
        ]

        # Write headers to Excel sheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Make headers bold
            cell.font = Font(bold=True)

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write data rows
        for row_num, ticket in enumerate(tickets, start=2):
            ws.cell(row=row_num, column=1, value=row_num-1)  # №
            ws.cell(row=row_num, column=2, value=ticket.ticket_number)
            ws.cell(row=row_num, column=3, value=ticket.ticket_sale.id)
            ws.cell(row=row_num, column=4, value=f"{ticket.event_date} {ticket.event_time}")
            ws.cell(row=row_num, column=5, value=ticket.amount)
            ws.cell(row=row_num, column=6, value=ticket.service.name)
            ws.cell(row=row_num, column=7, value=ticket.service.inventory.name if ticket.service.inventory else '')
            ws.cell(row=row_num, column=8, value=ticket.ticket_sale.get_sale_type_display())
            ws.cell(row=row_num, column=9, value=ticket.ticket_sale.booking_begin_date)
            ws.cell(row=row_num, column=10, value=ticket.payment_id)
            ws.cell(row=row_num, column=11, value=ticket.ticket_sale.phone)

            # Apply borders to each cell in the current row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

        total_summary = tickets.aggregate(
            total_amount=Sum('amount')
        )

        # Add total summary row
        summary_row_num = len(tickets) + 2  # Place the summary row after all data rows
        total_row = [
            'Итого',
            '', '', '', total_summary['total_amount'],
            '', '', '', '', '', ''
        ]
        for col_num, value in enumerate(total_row, 1):
            cell = ws.cell(row=summary_row_num, column=col_num, value=value)
            # Make the total row bold
            cell.font = Font(bold=True)
            # Apply borders to each cell in the total row
            cell.border = thin_border

        # Apply borders to the header row
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).border = thin_border

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create a response for the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="tickets_report.xlsx"'
        wb.save(response)

        return response
    else:
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')
        return redirect('reports:tickets_report')


@csrf_exempt
def ticket_print_data(request, ticket_id):
    tickets = TicketSalesTicket.objects.filter(id=ticket_id)
    print('>>>> ticket_id', ticket_id)
    data = {
        'services': [
            {
                'ticket_guid': str(ticket.ticket_guid),
                'number': ticket.number,
                'service_name': ticket.service.name,
                'event_date': ticket.event_date.strftime('%d.%m.%Y'),
                'event_time': f'{ticket.event_time.strftime('%H:%M')} - {ticket.event_time_end.strftime('%H:%M')}',
                'event_name': ticket.event.name,
                'amount': ticket.amount,
                'tickets_count': 1,
                'sale_id': ticket.ticket_sale.id,
            }
            for ticket in tickets
        ]
    }
    return JsonResponse(data)


@csrf_exempt
def ticket_print_pdf(request, ticket_id):
    tickets = TicketSalesTicket.objects.filter(id=ticket_id)

    if not tickets.exists():
        return HttpResponse("No ticket data available", status=404)

    # Generate the HTML content for the PDF
    html_content = ''
    for ticket in tickets:
        html_content += f'''
        <div style="text-align: center; font-family: 'Arial', sans-serif;">            
            <img src="https://api.qrserver.com/v1/create-qr-code/?size=210x210&data={ticket.ticket_guid}" alt="QR Code">
            <p>Билет № <strong>{ticket.ticket_sale.id}-{ticket.number}</strong></p>
            <p>Количество билетов: <strong>1</strong></p>
            <p><strong>Дата и время сеанса:</strong></p>
            <p>{ticket.event_date.strftime("%d.%m.%Y")} {ticket.event_time.strftime("%H:%M")} - {ticket.event_time_end.strftime("%H:%M")}</p>
            <p><strong>Услуга:</strong></p> 
            <p>{ticket.service.name}</p>
            <p><strong>Мероприятие:</strong></p>
            <p>{ticket.event.name}</p>
            <p><strong>Стоимость билета:</strong></p>
            <p>{ticket.amount} ₸</p>
            <p><strong>Подробная информация:</strong></p>
            <p>www.muzaidyny.kz тел: 8(7172) 242424</p>
        </div>        
        '''

    # from xhtml2pdf import pisa
    # # Use WeasyPrint to generate the
    # pdf = pisa.CreatePDF(html_content, dest=BytesIO())

    # Return the PDF as a response
    response = HttpResponse(html_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{ticket_id}.pdf"'

    return response


@permission_required('reports.view_services_report', raise_exception=True)
def services_report(request):
    form = ServiceReportForm(request.GET or None)

    if form.is_valid():
        report_data, services, dates = get_services_report_data(form)

        # Paginate services
        paginator = Paginator(list(services.keys()), 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Summary calculations (across all dates and services)
        summary = defaultdict(lambda: {'total_amount': 0, 'total_count': 0})

        for service_id in page_obj.object_list:
            if service_id in report_data:
                date_data = report_data[service_id]
                for date, data in date_data.items():
                    summary[date]['total_amount'] += data['amount']
                    summary[date]['total_count'] += data['count']

        # Преобразуем summary в OrderedDict, отсортированный по ключу (дате)
        summary = OrderedDict(sorted(summary.items()))
    else:
        page_obj = None
        report_data = {}
        dates = []
        services = {}
        summary = {}
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')

    context = {
        'form': form,
        'page_obj': page_obj,
        'report_data': report_data,
        'dates': dates,
        'services': services,
        'summary': summary,
    }

    return render(request, 'reports/service_report.html', context)


@permission_required('reports.view_services_report', raise_exception=True)
def services_report_excel_export(request):
    form = ServiceReportForm(request.GET or None)

    if not form.is_valid():
        messages.error(request, 'Пожалуйста исправьте ошибки в фильтрах')
        return redirect('reports:services_report')  # Замените на ваш URL-адрес

    report_data, services, dates = get_services_report_data(form)

    # Создаем Excel файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Services Report"

    # Стиль для заголовка и итогов
    bold_font = Font(bold=True)

    # Создаем первый и второй ряды заголовков
    date_headers = ['№', 'Наименование услуги']
    metric_headers = ['', '']

    for date in dates:
        date_headers.extend([date, None])  # Пара "дата, пустая ячейка"
        metric_headers.extend(["Количество", "Сумма"])  # Заголовки "количество, сумма"

    # Добавляем заголовки в два ряда
    sheet.append(date_headers)
    sheet.append(metric_headers)

    # Применяем стиль к заголовкам
    for cell in sheet[1] + sheet[2]:  # Первая и вторая строки - заголовки
        if cell:
            cell.font = bold_font

    # Объединяем ячейки дат с последующими пустыми ячейками
    col_index = 3  # Первая ячейка с датой находится в колонке 3
    for _ in dates:
        sheet.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + 1)
        col_index += 2  # Переходим к следующей паре ячеек

    # Готовим стиль границы ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заполняем строки данных
    row_number = 1  # Нумерация строк начинается с 1
    for service_id, service_name in services.items():
        row_number += 1  # Увеличиваем номер строки
        row = [row_number - 1, service_name]  # Добавляем номер строки и наименование услуги

        # Заполняем строки данных по датам
        for date in dates:
            # Количество билетов и сумма для каждой даты
            count = report_data[service_id][date]['count']
            amount = report_data[service_id][date]['amount']
            row.extend([count, amount])

        sheet.append(row)

    # Добавляем итоговую строку
    summary = defaultdict(lambda: {'total_amount': 0, 'total_count': 0})
    for service_id, date_data in report_data.items():
        for date, data in date_data.items():
            summary[date]['total_amount'] += data['amount']
            summary[date]['total_count'] += data['count']

    # Преобразуем summary в OrderedDict, отсортированный по ключу (дате)
    summary = OrderedDict(sorted(summary.items()))

    # Итоговая строка
    total_row = ['', 'Итого']
    for date in dates:
        total_row.append(summary[date]['total_count'])
        total_row.append(summary[date]['total_amount'])
    sheet.append(total_row)

    # Применяем стиль к итоговой строке
    for cell in sheet[sheet.max_row]:  # Последняя строка - итоги
        cell.font = bold_font

    # Применяем границы ко всем ячейкам
    for row in sheet.iter_rows():
        for cell in row:
            if cell:  # Убедиться, что ячейка не None
                cell.border = thin_border

    # Подготовка HTTP ответа
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=services_report.xlsx'

    # Сохраняем Excel файл в HTTP ответ
    workbook.save(response)
    return response
